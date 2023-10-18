# This script is used for taking the "tall" "all results flat file" PFAS sampling feature class and re-shaping it to be "wide."
# "Wide" means that each sample is one row in the table, with a unique field for each analyte's result and result qualifier. 

# This script assumes you've used the PFAS_ARFF_Tall script to get address IDs associated with samples. It also assumes you have the master PFAS address
# layer in your map. That layer is here: []

# Data should not just be "fed" to this script and considered good to go. There could be issues in the sampling data that need to be addressed.

# Add to this script if/when you do find things that need to be added so that the script can be made more robust.

# Last updated 10/11/2023

import pandas as pd # Used for re-shaping the table
import openpyxl # Used for working with Excel

# Set up some basic things

# Access "tall" "all results flat file" PFAS sampling feature class
aprx = arcpy.mp.ArcGISProject("CURRENT")
m = aprx.listMaps("Map3")[0]
addresses = m.listLayers("PFAS_Addresses")[0]
samplesFC = m.listLayers("Grayling_GAAF_SiteSummary_Copy_AllResultsFlatFile_XYEvent_FC")[0]

# Geodatabase location & where you want to save any Excel files
location = r"C:\Users\JohnsonN35\Local_Work\PFAS_Script"

# Name of the site (used for files)
site_ = "Grayling_GAAF"

# Excel extension
ext = ".xlsx"

# Geodatabase name (this script assumes your gdb is in the same place as where you want to save any Excel files)
gdb = "PFAS.gdb"

# Create a pandas dataframe

fields = ['Site','AddressID','Site_Name','Site_Subarea','Data_File_Name','Report_File_Name','Lab_Name','Lab_Work_Order','Lab_Sample_ID','Field_Sample_ID','Field_Location_Code','Duplicate','Sampled_Address_Clean','Sampling_Round','Sample_PrePost','Collect_Date','Collected_By','Matrix_Stdz','Analysis_Method_Stdz','Sample_NDE','Sample_TotalPFAS','Analyte_Abbrev','Result_Num','Result_Qualifier']

df = pd.DataFrame.from_records(data = arcpy.da.SearchCursor(samplesFC, fields), columns = fields)

df.head()

# Create a wide dataframe

# GRAYLING-SPECIFIC (leave out report_file_name since it contains invalid values)
df_wide = df.pivot(index = ['Site','AddressID','Site_Name','Site_Subarea','Data_File_Name','Lab_Name','Lab_Work_Order','Lab_Sample_ID','Field_Sample_ID','Field_Location_Code','Duplicate','Sampled_Address_Clean','Sampling_Round','Sample_PrePost','Collect_Date','Collected_By','Matrix_Stdz','Analysis_Method_Stdz','Sample_NDE','Sample_TotalPFAS'], columns = ['Analyte_Abbrev'], values = ['Result_Num', 'Result_Qualifier'])

# ALL OTHER SITES
# df_wide = df.pivot(index = ['Site','AddressID','Site_Name','Site_Subarea','Data_File_Name','Report_File_Name','Lab_Name','Lab_Work_Order','Lab_Sample_ID','Field_Sample_ID','Field_Location_Code','Duplicate','Sampled_Address_Clean','Sampling_Round','Sample_PrePost','Collect_Date','Collected_By','Matrix_Stdz','Analysis_Method_Stdz','Sample_NDE','Sample_TotalPFAS'], columns = ['Analyte_Abbrev'], values = ['Result_Num', 'Result_Qualifier'])

df_wide.reset_index(inplace = True)

df_wide.head()

# Why not include Analyte_NDE from the tall file? Only 7 analytes (as of October 2023) have these values, so it would be a lot of empty fields to deal with
# for not much gain. We'll create fields for these analytes' NDE status and re-calculate these values in a later step.

# Export to an Excel spreadsheet

name = site_ + "_" + "samplesFC_wide"

df_wide.to_excel(location + "/" + name + ext, sheet_name = 'Sheet1')

# Just do manually for Grayling: add a Report_File_Name field to the spreadsheet. It will remain empty.
# We're adding it because otherwise the column locations below would be off by 1.

# Clean up the Excel file (we can't use the current table structure in a feature class)

from openpyxl import load_workbook

outpath = location + "/" + name + ext
wb = load_workbook(outpath)
ws = wb.active

# Append all the result number fields with "_Result_Num"
# For Grayling GAAF, analyte result values go from cell W2 (23rd column) - cell CO2 (93rd column); these may be different for other sites, since
# the number of analyte columns is dependent on the number of analytes with results for a particular site.
for index, row in enumerate(ws.rows, start = 23):
        if index > 93: break
        ws.cell(row = 2, column = index).value += '_Result_Num'

# Append all the result qualifier fields with "_Result_Qualifier"
# For Grayling GAAF, analyte result qualifiers go from cell CP2 (94th column) - cell FH2 (164th column); these may be different for other sites, since
# the number of analyte columns is dependent on the number of analytes with results for a particular site.
for index, row in enumerate(ws.rows, start = 94):
        if index > 164: break
        ws.cell(row = 2, column = index).value += '_Result_Qualifier'

# Write the values of cells B1 (2nd column) - V1 (22nd column) into cells B2 - V2
for index, row in enumerate(ws.rows, start = 2):
        if index > 22: break
        ws.cell(row = 2, column = index).value = ws.cell(row = 1, column = index).value

# Unmerge cells (causes issues with deleting rows in next step)
# This might also need to be customized based on the site
ws.unmerge_cells('W1:CO1')
ws.unmerge_cells('CP1:FH1')

# Delete rows and columns we don't need / that can't be imported into a feature class
ws.delete_cols(1)
ws.delete_rows(1)
ws.delete_rows(2) # Once we delete the first row, the original third row becomes the second row
wb.save(outpath)

# Create a new table to hold the wide sampling results

tableName = site_ + "_Pivoted"

arcpy.management.CreateTable(location + "/" + gdb, tableName)

# Create the fields for the table; this includes ALL the PFAS analytes, including branched and linear. We probably don't need these, but they're there in case.
arcpy.management.AddFields(tableName, "\
Site TEXT 'Site' 255 # #;\
AddressID TEXT 'Address ID' 10 # #;\
Site_Name TEXT 'Site Name' 255 # #;\
Site_Subarea TEXT 'Site Subarea' 255 # #;\
Data_File_Name TEXT 'Data File Name' 255 # #;\
Report_File_Name TEXT 'Report File Name' 255 # #;\
Lab_Name TEXT 'Lab Name' 255 # #;\
Lab_Work_Order TEXT 'Lab Work Order' 255 # #;\
Lab_Sample_ID TEXT 'Lab Sample ID' 255 # #;\
Field_Sample_ID TEXT 'Field Sample ID' 255 # #;\
Field_Location_Code TEXT 'Field Location Code' 255 # #;\
Duplicate TEXT Duplicate 255 # #;\
Sampled_Address_Clean TEXT 'Sampled Address Clean' 255 # #;\
Sampling_Round TEXT 'Sampling Round' 255 # #;\
Sample_PrePost TEXT 'Sample PrePost' 255 # #;\
Collect_Date DATE 'Collect Date' # # #;\
Collected_By TEXT 'Collected By' 255 # #;\
Matrix_Stdz TEXT 'Matrix' 255 # #;\
Analysis_Method_Stdz TEXT 'Analysis Method' 255 # #;\
Sample_NDE TEXT 'Sample NDE' 2 # #;\
Sample_TotalPFAS DOUBLE 'Sample Total PFAS' # # #;\
HFPO_DA_GenX_NDE TEXT 'HFPO-DA (GenX) NDE' 2 # #;\
HFPO_DA_GenX_Result_Num DOUBLE 'HFPO-DA (GenX) Result Num' # # #;\
HFPO_DA_GenX_Result_Qualifier TEXT 'HFPO-DA (GenX) Result Qualifier' 20 # #;\
PFBS_NDE TEXT 'PFBS NDE' 2 # #;\
PFBS_Result_Num DOUBLE 'PFBS Result Num' # # #;\
PFBS_Result_Qualifier TEXT 'PFBS Result Qualifier' 20 # #;\
PFHxA_NDE TEXT 'PFHxA NDE' 2 # #;\
PFHxA_Result_Num DOUBLE 'PFHxA Result Num' # # #;\
PFHxA_Result_Qualifier TEXT 'PFHxA Result Qualifier' 20 # #;\
PFHxS_NDE TEXT 'PFHxS NDE' 2 # #;\
PFHxS_Result_Num DOUBLE 'PFHxS Result Num' # # #;\
PFHxS_Result_Qualifier TEXT 'PFHxS Result Qualifier' 20 # #;\
PFOA_NDE TEXT 'PFOA NDE' 2 # #;\
PFOA_Result_Num DOUBLE 'PFOA Result Num' # # #;\
PFOA_Result_Qualifier TEXT 'PFOA Result Qualifier' 20 # #;\
PFOS_NDE TEXT 'PFOS NDE' 2 # #;\
PFOS_Result_Num DOUBLE 'PFOS Result Num' # # #;\
PFOS_Result_Qualifier TEXT 'PFOS Result Qualifier' 20 # #;\
PFNA_NDE TEXT 'PFNA NDE' 2 # #;\
PFNA_Result_Num DOUBLE 'PFNA Result Num' # # #;\
PFNA_Result_Qualifier TEXT 'PFNA Result Qualifier' 20 # #;\
ADONA_Result_Num DOUBLE 'ADONA Result Num' # # #;\
ADONA_Result_Qualifier TEXT 'ADONA Result Qualifier' 20 # #;\
Br_PFHxS_Result_Num DOUBLE 'Br-PFHxS Result Num' # # #;\
Br_PFHxS_Result_Qualifier TEXT 'Br-PFHxS Result Qualifier' 20 # #;\
Br_PFOA_Result_Num DOUBLE 'Br-PFOA Result Num' # # #;\
Br_PFOA_Result_Qualifier TEXT 'Br-PFOA Result Qualifier' 20 # #;\
Br_PFOS_Result_Num DOUBLE 'Br-PFOS Result Num' # # #;\
Br_PFOS_Result_Qualifier TEXT 'Br-PFOS Result Qualifier' 20 # #;\
EtFOSA_Result_Num DOUBLE 'EtFOSA Result Num' # # #;\
EtFOSA_Result_Qualifier TEXT 'EtFOSA Result Qualifier' 20 # #;\
EtFOSAA_Result_Num DOUBLE 'EtFOSAA Result Num' # # #;\
EtFOSAA_Result_Qualifier TEXT 'EtFOSAA Result Qualifier' 20 # #;\
EtFOSE_Result_Num DOUBLE 'EtFOSE Result Num' # # #;\
EtFOSE_Result_Qualifier TEXT 'EtFOSE Result Qualifier' 20 # #;\
F_11Cl_PF3OUdS_Result_Num DOUBLE '11Cl-PF3OUdS Result Num' # # #;\
F_11Cl_PF3OUdS_Result_Qualifier TEXT '11Cl-PF3OUdS Result Qualifier' 20 # #;\
F_3_3FTCA_Result_Num DOUBLE '3:3 FTCA Result Num' # # #;\
F_3_3FTCA_Result_Qualifier TEXT '3:3 FTCA Result Qualifier' 20 # #;\
F_4_2FTS_Result_Num DOUBLE '4:2 FTS Result Num' # # #;\
F_4_2FTS_Result_Qualifier TEXT '4:2 FTS Result Qualifier' 20 # #;\
F_5_3FTCA_Result_Num DOUBLE '5:3 FTCA Result Num' # # #;\
F_5_3FTCA_Result_Qualifier TEXT '5:3 FTCA Result Qualifier' 20 # #;\
F_6_2FTS_Result_Num DOUBLE '6:2 FTS Result Num' # # #;\
F_6_2FTS_Result_Qualifier TEXT '6:2 FTS Result Qualifier' 20 # #;\
F_7_3FTCA_Result_Num DOUBLE '7:3 FTCA Result Num' # # #;\
F_7_3FTCA_Result_Qualifier TEXT '7:3 FTCA Result Qualifier' 20 # #;\
F_8_2FTS_Result_Num DOUBLE '8:2 FTS Result Num' # # #;\
F_8_2FTS_Result_Qualifier TEXT '8:2 FTS Result Qualifier' 20 # #;\
F_9Cl_PF3ONS_Result_Num DOUBLE '9Cl-PF3ONS Result Num' # # #;\
F_9Cl_PF3ONS_Result_Qualifier TEXT '9Cl-PF3ONS Result Qualifier' 20 # #;\
L_6_2FTS_Result_Num DOUBLE 'L-6:2 FTS Result Num' # # #;\
L_6_2FTS_Result_Qualifier TEXT 'L-6:2 FTS Result Qualifier' 20 # #;\
L_8_2FTS_Result_Num DOUBLE 'L-8:2 FTS Result Num' # # #;\
L_8_2FTS_Result_Qualifier TEXT 'L-8:2FTS Result Qualifier' 20 # #;\
L_PFBA_Result_Num DOUBLE 'L-PFBA Result Num' # # #;\
L_PFBA_Result_Qualifier TEXT 'L-PFBA Result Qualifier' 20 # #;\
L_PFBS_Result_Num DOUBLE 'L-PFBS Result Num' # # #;\
L_PFBS_Result_Qualifier TEXT 'L-PFBS Result Qualifier' 20 # #;\
L_PFDA_Result_Num DOUBLE 'L-PFDA Result Num' # # #;\
L_PFDA_Result_Qualifier TEXT 'L-PFDA Result Qualifier' 20 # #;\
L_PFDoDA_Result_Num DOUBLE 'L-PFDoDA Result Num' # # #;\
L_PFDoDA_Result_Qualifier TEXT 'L-PFDoDA Result Qualifier' 20 # #;\
L_PFDS_Result_Num DOUBLE 'L-PFDS Result Num' # # #;\
L_PFDS_Result_Qualifier TEXT 'L-PFDS Result Qualifier' 20 # #;\
L_PFHpA_Result_Num DOUBLE 'L-PFHpA Result Num' # # #;\
L_PFHpA_Result_Qualifier TEXT 'L-PFHpA Result Qualifier' 20 # #;\
L_PFHpS_Result_Num DOUBLE 'L-PFHpS Result Num' # # #;\
L_PFHpS_Result_Qualifier TEXT 'L-PFHpS Result Qualifier' 20 # #;\
L_PFHxA_Result_Num DOUBLE 'L-PFHxA Result Num' # # #;\
L_PFHxA_Result_Qualifier TEXT 'L-PFHxA Result Qualifier' 20 # #;\
L_PFHxDA_Result_Num DOUBLE 'L-PFHxDA Result Num' # # #;\
L_PFHxDA_Result_Qualifier TEXT 'L-PFHxDA Result Qualifier' 20 # #;\
L_PFHxS_Result_Num DOUBLE 'L-PFHxS Result Num' # # #;\
L_PFHxS_Result_Qualifier TEXT 'L-PFHxS Result Qualifier' 20 # #;\
L_PFNA_Result_Num DOUBLE 'L-PFNA Result Num' # # #;\
L_PFNA_Result_Qualifier TEXT 'L-PFNA Result Qualifier' 20 # #;\
L_PFOA_Result_Num DOUBLE 'L-PFOA Result Num' # # #;\
L_PFOA_Result_Qualifier TEXT 'L-PFOA Result Qualifier' 20 # #;\
L_PFODA_Result_Num DOUBLE 'L-PFODA Result Num' # # #;\
L_PFODA_Result_Qualifier TEXT 'L-PFODA Result Qualifier' 20 # #;\
L_PFOS_Result_Num DOUBLE 'L-PFOS Result Num' # # #;\
L_PFOS_Result_Qualifier TEXT 'L-PFOS Result Qualifier' 20 # #;\
L_PFOSA_Result_Num DOUBLE 'L-PFOSA Result Num' # # #;\
L_PFOSA_Result_Qualifier TEXT 'L-PFOSA Result Qualifier' 20 # #;\
L_PFPeA_Result_Num DOUBLE 'L-PFPeA Result Num' # # #;\
L_PFPeA_Result_Qualifier TEXT 'L-PFPeA Result Qualifier' 20 # #;\
L_PFTeDA_Result_Num DOUBLE 'L-PFTeDA Result Num' # # #;\
L_PFTeDA_Result_Qualifier TEXT 'L-PFTeDA Result Qualifier' 20 # #;\
L_PFTrDA_Result_Num DOUBLE 'L-PFTrDA Result Num' # # #;\
L_PFTrDA_Result_Qualifier TEXT 'L-PFTrDA Result Qualifier' 20 # #;\
L_PFUnDA_Result_Num DOUBLE 'L-PFUnDA Result Num' # # #;\
L_PFUnDA_Result_Qualifier TEXT 'L-PFUnDA Result Qualifier' 20 # #;\
MeFOSA_Result_Num DOUBLE 'MeFOSA Result Num' # # #;\
MeFOSA_Result_Qualifier TEXT 'MeFOSA Result Qualifier' 20 # #;\
MeFOSAA_Result_Num DOUBLE 'MeFOSAA Result Num' # # #;\
MeFOSAA_Result_Qualifier TEXT 'MeFOSAA Result Qualifier' 20 # #;\
MeFOSE_Result_Num DOUBLE 'MeFOSE Result Num' # # #;\
MeFOSE_Result_Qualifier TEXT 'MeFOSE Result Qualifier' 20 # #;\
NFDHA_Result_Num DOUBLE 'NFDHA Result Num' # # #;\
NFDHA_Result_Qualifier TEXT 'NFDHA Result Qualifier' 20 # #;\
PFBA_Result_Num DOUBLE 'PFBA Result Num' # # #;\
PFBA_Result_Qualifier TEXT 'PFBA Result Qualifier' 20 # #;\
PFBSA_Result_Num DOUBLE 'PFBSA Result Num' # # #;\
PFBSA_Result_Qualifier TEXT 'PFBSA Result Qualifier' 20 # #;\
PFDA_Result_Num DOUBLE 'PFDA Result Num' # # #;\
PFDA_Result_Qualifier TEXT 'PFDA Result Qualifier' 20 # #;\
PFDoDA_Result_Num DOUBLE 'PFDoDA Result Num' # # #;\
PFDoDA_Result_Qualifier TEXT 'PFDoDA Result Qualifier' 20 # #;\
PFDS_Result_Num DOUBLE 'PFDS Result Num' # # #;\
PFDS_Result_Qualifier TEXT 'PFDS Result Qualifier' 20 # #;\
PFECHS_Result_Num DOUBLE 'PFECHS Result Num' # # #;\
PFECHS_Result_Qualifier TEXT 'PFECHS Result Qualifier' 20 # #;\
PFEESA_Result_Num DOUBLE 'PFEESA Result Num' # # #;\
PFEESA_Result_Qualifier TEXT 'PFEESA Result Qualifier' 20 # #;\
PFHpA_Result_Num DOUBLE 'PFHpA Result Num' # # #;\
PFHpA_Result_Qualifier TEXT 'PFHpA Result Qualifier' 20 # #;\
PFHpS_Result_Num DOUBLE 'PFHpS Result Num' # # #;\
PFHpS_Result_Qualifier TEXT 'PFHpS Result Qualifier' 20 # #;\
PFHxDA_Result_Num DOUBLE 'PFHxDA Result Num' # # #;\
PFHxDA_Result_Qualifier TEXT 'PFHxDA Result Qualifier' 20 # #;\
PFHxSA_Result_Num DOUBLE 'PFHxSA Result Num' # # #;\
PFHxSA_Result_Qualifier TEXT 'PFHxSA Result Qualifier' 20 # #;\
PFMBA_Result_Num DOUBLE 'PFMBA Result Num' # # #;\
PFMBA_Result_Qualifier TEXT 'PFMBA Result Qualifier' 20 # #;\
PFMPA_Result_Num DOUBLE 'PFMPA Result Num' # # #;\
PFMPA_Result_Qualifier TEXT 'PFMPA Result Qualifier' 20 # #;\
PFNS_Result_Num DOUBLE 'PFNS Result Num' # # #;\
PFNS_Result_Qualifier TEXT 'PFNS Result Qualifier' 20 # #;\
PFOA_PFOS_Result_Num DOUBLE 'PFOA+PFOS Result Num' # # #;\
PFOA_PFOS_Result_Qualifier TEXT 'PFOA+PFOS Result Qualifier' 20 # #;\
PFODA_Result_Num DOUBLE 'PFODA Result Num' # # #;\
PFODA_Result_Qualifier TEXT 'PFODA Result Qualifier' 20 # #;\
PFOSA_Result_Num DOUBLE 'PFOSA Result Num' # # #;\
PFOSA_Result_Qualifier TEXT 'PFOSA Result Qualifier' 20 # #;\
PFPeA_Result_Num DOUBLE 'PFPeA Result Num' # # #;\
PFPeA_Result_Qualifier TEXT 'PFPeA Result Qualifier' 20 # #;\
PFPeS_Result_Num DOUBLE 'PFPeS Result Num' # # #;\
PFPeS_Result_Qualifier TEXT 'PFPeS Result Qualifier' 20 # #;\
PFPrS_Result_Num DOUBLE 'PFPrS Result Num' # # #;\
PFPrS_Result_Qualifier TEXT 'PFPrS Result Qualifier' 20 # #;\
PFTeDA_Result_Num DOUBLE 'PFTeDA Result Num' # # #;\
PFTeDA_Result_Qualifier TEXT 'PFTeDA Result Qualifier' 20 # #;\
PFTrDA_Result_Num DOUBLE 'PFTrDA Result Num' # # #;\
PFTrDA_Result_Qualifier TEXT 'PFTrDA Result Qualifier' 20 # #;\
PFUnDA_Result_Num DOUBLE 'PFUnDA Result Num' # # #;\
PFUnDA_Result_Qualifier TEXT 'PFUnDA Result Qualifier' 20 # #;\
TotalPFAS_Result_Num DOUBLE 'Total PFAS Result Num' # # #;\
TotalPFAS_Result_Qualifier TEXT 'Total PFAS Result Qualifier' 20 # #;\
")

# Append data to the table

inputTable = location + "/" + name + ext + "/" + "Sheet1$"
targetTable = m.listTables(tableName)[0]

field_mapping = '\
Site "Site" true true false 255 Text 0 0,First,#,inputTable,Site,0,255;\
AddressID "Address ID" true true false 10 Text 0 0,First,#,inputTable,AddressID,0,255;\
Site_Name "Site Name" true true false 255 Text 0 0,First,#,inputTable,Site_Name,0,255;\
Site_Subarea "Site Subarea" true true false 255 Text 0 0,First,#,inputTable,Site_Subarea,0,255;\
Data_File_Name "Data File Name" true true false 255 Text 0 0,First,#,inputTable,Data_File_Name,0,255;\
Report_File_Name "Report File Name" true true false 255 Text 0 0,First,#,inputTable,Report_File_Name,0,255;\
Lab_Name "Lab Name" true true false 255 Text 0 0,First,#,inputTable,Lab_Name,0,255;\
Lab_Work_Order "Lab Work Order" true true false 255 Text 0 0,First,#,inputTable,Lab_Work_Order,0,255;\
Lab_Sample_ID "Lab Sample ID" true true false 255 Text 0 0,First,#,inputTable,Lab_Sample_ID,0,255;\
Field_Sample_ID "Field Sample ID" true true false 255 Text 0 0,First,#,inputTable,Field_Sample_ID,0,255;\
Field_Location_Code "Field Location Code" true true false 255 Text 0 0,First,#,inputTable,Field_Location_Code,0,255;\
Duplicate "Duplicate" true true false 255 Text 0 0,First,#,inputTable,Duplicate,0,255;\
Sampled_Address_Clean "Sampled Address Clean" true true false 255 Text 0 0,First,#,inputTable,Sampled_Address_Clean,0,255;\
Sampling_Round "Sampling Round" true true false 255 Text 0 0,First,#,inputTable,Sampling_Round,0,255;\
Sample_PrePost "Sample PrePost" true true false 255 Text 0 0,First,#,inputTable,Sample_PrePost,0,255;\
Collect_Date "Collect Date" true true false 8 Date 0 0,First,#,inputTable,Collect_Date,-1,-1;\
Collected_By "Collected By" true true false 255 Text 0 0,First,#,inputTable,Collected_By,0,255;\
Matrix_Stdz "Matrix" true true false 255 Text 0 0,First,#,inputTable,Matrix_Stdz,0,255;\
Analysis_Method_Stdz "Analysis Method" true true false 255 Text 0 0,First,#,inputTable,Analysis_Method_Stdz,0,255;\
Sample_NDE "Sample NDE" true true false 2 Text 0 0,First,#,inputTable,Sample_NDE,0,255;\
Sample_TotalPFAS "Sample Total PFAS" true true false 8 Double 0 0,First,#,inputTable,Sample_TotalPFAS,-1,-1;\
HFPO_DA_GenX_NDE "HFPO-DA (GenX) NDE" true true false 2 Text 0 0,First,#;\
HFPO_DA_GenX_Result_Num "HFPO-DA (GenX) Result Num" true true false 8 Double 0 0,First,#,inputTable,HFPO_DA__GenX__Result_Num,-1,-1;\
HFPO_DA_GenX_Result_Qualifier "HFPO-DA (GenX) Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,HFPO_DA__GenX__Result_Qualifier,0,255;\
PFBS_NDE "PFBS NDE" true true false 2 Text 0 0,First,#;\
PFBS_Result_Num "PFBS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFBS_Result_Num,-1,-1;\
PFBS_Result_Qualifier "PFBS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFBS_Result_Qualifier,0,255;\
PFHxA_NDE "PFHxA NDE" true true false 2 Text 0 0,First,#;\
PFHxA_Result_Num "PFHxA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFHxA_Result_Num,-1,-1;\
PFHxA_Result_Qualifier "PFHxA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFHxA_Result_Qualifier,0,255;\
PFHxS_NDE "PFHxS NDE" true true false 2 Text 0 0,First,#;\
PFHxS_Result_Num "PFHxS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFHxS_Result_Num,-1,-1;\
PFHxS_Result_Qualifier "PFHxS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFHxS_Result_Qualifier,0,255;\
PFOA_NDE "PFOA NDE" true true false 2 Text 0 0,First,#;\
PFOA_Result_Num "PFOA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFOA_Result_Num,-1,-1;\
PFOA_Result_Qualifier "PFOA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFOA_Result_Qualifier,0,255;\
PFOS_NDE "PFOS NDE" true true false 2 Text 0 0,First,#;\
PFOS_Result_Num "PFOS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFOS_Result_Num,-1,-1;\
PFOS_Result_Qualifier "PFOS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFOS_Result_Qualifier,0,255;\
PFNA_NDE "PFNA NDE" true true false 2 Text 0 0,First,#;\
PFNA_Result_Num "PFNA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFNA_Result_Num,-1,-1;\
PFNA_Result_Qualifier "PFNA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFNA_Result_Qualifier,0,255;\
ADONA_Result_Num "ADONA Result Num" true true false 8 Double 0 0,First,#,inputTable,ADONA_Result_Num,-1,-1;\
ADONA_Result_Qualifier "ADONA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,ADONA_Result_Qualifier,0,255;\
Br_PFHxS_Result_Num "Br-PFHxS Result Num" true true false 8 Double 0 0,First,#,inputTable,Br_PFHxS_Result_Num,-1,-1;\
Br_PFHxS_Result_Qualifier "Br-PFHxS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,Br_PFHxS_Result_Qualifier,0,255;\
Br_PFOA_Result_Num "Br-PFOA Result Num" true true false 8 Double 0 0,First,#,inputTable,Br_PFOA_Result_Num,-1,-1;\
Br_PFOA_Result_Qualifier "Br-PFOA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,Br_PFOA_Result_Qualifier,0,255;\
Br_PFOS_Result_Num "Br-PFOS Result Num" true true false 8 Double 0 0,First,#,inputTable,Br_PFOS_Result_Num,-1,-1;\
Br_PFOS_Result_Qualifier "Br-PFOS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,Br_PFOS_Result_Qualifier,0,255;\
EtFOSA_Result_Num "EtFOSA Result Num" true true false 8 Double 0 0,First,#,inputTable,EtFOSA_Result_Num,-1,-1;\
EtFOSA_Result_Qualifier "EtFOSA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,EtFOSA_Result_Qualifier,0,255;\
EtFOSAA_Result_Num "EtFOSAA Result Num" true true false 8 Double 0 0,First,#,inputTable,EtFOSAA_Result_Num,-1,-1;\
EtFOSAA_Result_Qualifier "EtFOSAA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,EtFOSAA_Result_Qualifier,0,255;\
EtFOSE_Result_Num "EtFOSE Result Num" true true false 8 Double 0 0,First,#,inputTable,EtFOSE_Result_Num,-1,-1;\
EtFOSE_Result_Qualifier "EtFOSE Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,EtFOSE_Result_Qualifier,0,255;\
F_11Cl_PF3OUdS_Result_Num "11Cl-PF3OUdS Result Num" true true false 8 Double 0 0,First,#,inputTable,F11Cl_PF3OUdS_Result_Num,-1,-1;\
F_11Cl_PF3OUdS_Result_Qualifier "11Cl-PF3OUdS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,F11Cl_PF3OUdS_Result_Qualifier,0,255;\
F_3_3FTCA_Result_Num "3:3 FTCA Result Num" true true false 8 Double 0 0,First,#,inputTable,F3_3_FTCA_Result_Num,-1,-1;\
F_3_3FTCA_Result_Qualifier "3:3 FTCA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,F3_3_FTCA_Result_Qualifier,0,255;\
F_4_2FTS_Result_Num "4:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputTable,F4_2_FTS_Result_Num,-1,-1;\
F_4_2FTS_Result_Qualifier "4:2 FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,F4_2_FTS_Result_Qualifier,0,255;\
F_5_3FTCA_Result_Num "5:3 FTCA Result Num" true true false 8 Double 0 0,First,#,inputTable,F5_3_FTCA_Result_Num,-1,-1;\
F_5_3FTCA_Result_Qualifier "5:3 FTCA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,F5_3_FTCA_Result_Qualifier,0,255;\
F_6_2FTS_Result_Num "6:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputTable,F6_2_FTS_Result_Num,-1,-1;\
F_6_2FTS_Result_Qualifier "6:2 FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,F6_2_FTS_Result_Qualifier,0,255;\
F_7_3FTCA_Result_Num "7:3 FTCA Result Num" true true false 8 Double 0 0,First,#,inputTable,F7_3_FTCA_Result_Num,-1,-1;\
F_7_3FTCA_Result_Qualifier "7:3 FTCA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,F7_3_FTCA_Result_Qualifier,0,255;\
F_8_2FTS_Result_Num "8:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputTable,F8_2_FTS_Result_Num,-1,-1;\
F_8_2FTS_Result_Qualifier "8:2 FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,F8_2_FTS_Result_Qualifier,0,255;\
F_9Cl_PF3ONS_Result_Num "9Cl-PF3ONS Result Num" true true false 8 Double 0 0,First,#,inputTable,F9Cl_PF3ONS_Result_Num,-1,-1;\
F_9Cl_PF3ONS_Result_Qualifier "9Cl-PF3ONS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,F9Cl_PF3ONS_Result_Qualifier,0,255;\
L_6_2FTS_Result_Num "L-6:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputTable,L_6_2_FTS_Result_Num,-1,-1;\
L_6_2FTS_Result_Qualifier "L-6:2 FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_6_2_FTS_Result_Qualifier,0,255;\
L_8_2FTS_Result_Num "L-8:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputTable,L_8_2FTS_Result_Num,-1,-1;\
L_8_2FTS_Result_Qualifier "L-8:2FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_8_2FTS_Result_Qualifier,0,255;\
L_PFBA_Result_Num "L-PFBA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFBA_Result_Num,-1,-1;\
L_PFBA_Result_Qualifier "L-PFBA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFBA_Result_Qualifier,0,255;\
L_PFBS_Result_Num "L-PFBS Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFBS_Result_Num,-1,-1;\
L_PFBS_Result_Qualifier "L-PFBS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFBS_Result_Qualifier,0,255;\
L_PFDA_Result_Num "L-PFDA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFDA_Result_Num,-1,-1;\
L_PFDA_Result_Qualifier "L-PFDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFDA_Result_Qualifier,0,255;\
L_PFDoDA_Result_Num "L-PFDoDA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFDoDA_Result_Num,-1,-1;\
L_PFDoDA_Result_Qualifier "L-PFDoDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFDoDA_Result_Qualifier,0,255;\
L_PFDS_Result_Num "L-PFDS Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFDS_Result_Num,-1,-1;\
L_PFDS_Result_Qualifier "L-PFDS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFDS_Result_Qualifier,0,255;\
L_PFHpA_Result_Num "L-PFHpA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFHpA_Result_Num,-1,-1;\
L_PFHpA_Result_Qualifier "L-PFHpA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFHpA_Result_Qualifier,0,255;\
L_PFHpS_Result_Num "L-PFHpS Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFHpS_Result_Num,-1,-1;\
L_PFHpS_Result_Qualifier "L-PFHpS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFHpS_Result_Qualifier,0,255;\
L_PFHxA_Result_Num "L-PFHxA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFHxA_Result_Num,-1,-1;\
L_PFHxA_Result_Qualifier "L-PFHxA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFHxA_Result_Qualifier,0,255;\
L_PFHxDA_Result_Num "L-PFHxDA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFHxDA_Result_Num,-1,-1;\
L_PFHxDA_Result_Qualifier "L-PFHxDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFHxDA_Result_Qualifier,0,255;\
L_PFHxS_Result_Num "L-PFHxS Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFHxS_Result_Num,-1,-1;\
L_PFHxS_Result_Qualifier "L-PFHxS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFHxS_Result_Qualifier,0,255;\
L_PFNA_Result_Num "L-PFNA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFNA_Result_Num,-1,-1;\
L_PFNA_Result_Qualifier "L-PFNA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFNA_Result_Qualifier,0,255;\
L_PFOA_Result_Num "L-PFOA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFOA_Result_Num,-1,-1;\
L_PFOA_Result_Qualifier "L-PFOA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFOA_Result_Qualifier,0,255;\
L_PFODA_Result_Num "L-PFODA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFODA_Result_Num,-1,-1;\
L_PFODA_Result_Qualifier "L-PFODA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFODA_Result_Qualifier,0,255;\
L_PFOS_Result_Num "L-PFOS Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFOS_Result_Num,-1,-1;\
L_PFOS_Result_Qualifier "L-PFOS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFOS_Result_Qualifier,0,255;\
L_PFOSA_Result_Num "L-PFOSA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFOSA_Result_Num,-1,-1;\
L_PFOSA_Result_Qualifier "L-PFOSA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFOSA_Result_Qualifier,0,255;\
L_PFPeA_Result_Num "L-PFPeA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFPeA_Result_Num,-1,-1;\
L_PFPeA_Result_Qualifier "L-PFPeA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFPeA_Result_Qualifier,0,255;\
L_PFTeDA_Result_Num "L-PFTeDA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFTeDA_Result_Num,-1,-1;\
L_PFTeDA_Result_Qualifier "L-PFTeDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFTeDA_Result_Qualifier,0,255;\
L_PFTrDA_Result_Num "L-PFTrDA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFTrDA_Result_Num,-1,-1;\
L_PFTrDA_Result_Qualifier "L-PFTrDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFTrDA_Result_Qualifier,0,255;\
L_PFUnDA_Result_Num "L-PFUnDA Result Num" true true false 8 Double 0 0,First,#,inputTable,L_PFUnDA_Result_Num,-1,-1;\
L_PFUnDA_Result_Qualifier "L-PFUnDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,L_PFUnDA_Result_Qualifier,0,255;\
MeFOSA_Result_Num "MeFOSA Result Num" true true false 8 Double 0 0,First,#,inputTable,MeFOSA_Result_Num,-1,-1;\
MeFOSA_Result_Qualifier "MeFOSA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,MeFOSA_Result_Qualifier,0,255;\
MeFOSAA_Result_Num "MeFOSAA Result Num" true true false 8 Double 0 0,First,#,inputTable,MeFOSAA_Result_Num,-1,-1;\
MeFOSAA_Result_Qualifier "MeFOSAA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,MeFOSAA_Result_Qualifier,0,255;\
MeFOSE_Result_Num "MeFOSE Result Num" true true false 8 Double 0 0,First,#,inputTable,MeFOSE_Result_Num,-1,-1;\
MeFOSE_Result_Qualifier "MeFOSE Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,MeFOSE_Result_Qualifier,0,255;\
NFDHA_Result_Num "NFDHA Result Num" true true false 8 Double 0 0,First,#,inputTable,NFDHA_Result_Num,-1,-1;\
NFDHA_Result_Qualifier "NFDHA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,NFDHA_Result_Qualifier,0,255;\
PFBA_Result_Num "PFBA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFBA_Result_Num,-1,-1;\
PFBA_Result_Qualifier "PFBA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFBA_Result_Qualifier,0,255;\
PFBSA_Result_Num "PFBSA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFBSA_Result_Num,-1,-1;\
PFBSA_Result_Qualifier "PFBSA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFBSA_Result_Qualifier,0,255;\
PFDA_Result_Num "PFDA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFDA_Result_Num,-1,-1;\
PFDA_Result_Qualifier "PFDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFDA_Result_Qualifier,0,255;\
PFDoDA_Result_Num "PFDoDA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFDoDA_Result_Num,-1,-1;\
PFDoDA_Result_Qualifier "PFDoDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFDoDA_Result_Qualifier,0,255;\
PFDS_Result_Num "PFDS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFDS_Result_Num,-1,-1;\
PFDS_Result_Qualifier "PFDS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFDS_Result_Qualifier,0,255;\
PFECHS_Result_Num "PFECHS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFECHS_Result_Num,-1,-1;\
PFECHS_Result_Qualifier "PFECHS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFECHS_Result_Qualifier,0,255;\
PFEESA_Result_Num "PFEESA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFEESA_Result_Num,-1,-1;\
PFEESA_Result_Qualifier "PFEESA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFEESA_Result_Qualifier,0,255;\
PFHpA_Result_Num "PFHpA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFHpA_Result_Num,-1,-1;\
PFHpA_Result_Qualifier "PFHpA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFHpA_Result_Qualifier,0,255;\
PFHpS_Result_Num "PFHpS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFHpS_Result_Num,-1,-1;\
PFHpS_Result_Qualifier "PFHpS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFHpS_Result_Qualifier,0,255;\
PFHxDA_Result_Num "PFHxDA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFHxDA_Result_Num,-1,-1;\
PFHxDA_Result_Qualifier "PFHxDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFHxDA_Result_Qualifier,0,255;\
PFHxSA_Result_Num "PFHxSA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFHxSA_Result_Num,-1,-1;\
PFHxSA_Result_Qualifier "PFHxSA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFHxSA_Result_Qualifier,0,255;\
PFMBA_Result_Num "PFMBA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFMBA_Result_Num,-1,-1;\
PFMBA_Result_Qualifier "PFMBA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFMBA_Result_Qualifier,0,255;\
PFMPA_Result_Num "PFMPA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFMPA_Result_Num,-1,-1;\
PFMPA_Result_Qualifier "PFMPA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFMPA_Result_Qualifier,0,255;\
PFNS_Result_Num "PFNS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFNS_Result_Num,-1,-1;\
PFNS_Result_Qualifier "PFNS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFNS_Result_Qualifier,0,255;\
PFOA_PFOS_Result_Num "PFOA+PFOS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFOA_PFOS_Result_Num,-1,-1;\
PFOA_PFOS_Result_Qualifier "PFOA+PFOS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFOA_PFOS_Result_Qualifier,0,255;\
PFODA_Result_Num "PFODA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFODA_Result_Num,-1,-1;\
PFODA_Result_Qualifier "PFODA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFODA_Result_Qualifier,0,255;\
PFOSA_Result_Num "PFOSA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFOSA_Result_Num,-1,-1;\
PFOSA_Result_Qualifier "PFOSA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFOSA_Result_Qualifier,0,255;\
PFPeA_Result_Num "PFPeA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFPeA_Result_Num,-1,-1;\
PFPeA_Result_Qualifier "PFPeA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFPeA_Result_Qualifier,0,255;\
PFPeS_Result_Num "PFPeS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFPeS_Result_Num,-1,-1;\
PFPeS_Result_Qualifier "PFPeS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFPeS_Result_Qualifier,0,255;\
PFPrS_Result_Num "PFPrS Result Num" true true false 8 Double 0 0,First,#,inputTable,PFPrS_Result_Num,-1,-1;\
PFPrS_Result_Qualifier "PFPrS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFPrS_Result_Qualifier,0,255;\
PFTeDA_Result_Num "PFTeDA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFTeDA_Result_Num,-1,-1;\
PFTeDA_Result_Qualifier "PFTeDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFTeDA_Result_Qualifier,0,255;\
PFTrDA_Result_Num "PFTrDA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFTrDA_Result_Num,-1,-1;\
PFTrDA_Result_Qualifier "PFTrDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFTrDA_Result_Qualifier,0,255;\
PFUnDA_Result_Num "PFUnDA Result Num" true true false 8 Double 0 0,First,#,inputTable,PFUnDA_Result_Num,-1,-1;\
PFUnDA_Result_Qualifier "PFUnDA Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,PFUnDA_Result_Qualifier,0,255;\
TotalPFAS_Result_Num "Total PFAS Result Num" true true false 8 Double 0 0,First,#,inputTable,Total_PFAS_Result_Num,-1,-1;\
TotalPFAS_Result_Qualifier "Total PFAS Result Qualifier" true true false 20 Text 0 0,First,#,inputTable,Total_PFAS_Result_Qualifier,0,255\
'

arcpy.management.Append(inputTable, targetTable, "NO_TEST", field_mapping)

# Calculate NDEs

arcpy.management.SelectLayerByAttribute(targetTable, "CLEAR_SELECTION", "CLEAR_SELECTION")

# HFPO-DA (GenX)
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "HFPO_DA_GenX_Result_Num = 0")
arcpy.management.CalculateField(targetTable, "HFPO_DA_GenX_NDE", '"ND"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "HFPO_DA_GenX_Result_Num > 0 And HFPO_DA_GenX_Result_Num <= 370")
arcpy.management.CalculateField(targetTable, "HFPO_DA_GenX_NDE", '"D"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "HFPO_DA_GenX_Result_Num > 370")
arcpy.management.CalculateField(targetTable, "HFPO_DA_GenX_NDE", '"E"')


# PFBS
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFBS_Result_Num = 0")
arcpy.management.CalculateField(targetTable, "PFBS_NDE", '"ND"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFBS_Result_Num > 0 And PFBS_Result_Num <= 420")
arcpy.management.CalculateField(targetTable, "PFBS_NDE", '"D"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFBS_Result_Num > 420")
arcpy.management.CalculateField(targetTable, "PFBS_NDE", '"E"')


# PFHxA
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFHxA_Result_Num = 0")
arcpy.management.CalculateField(targetTable, "PFHxA_NDE", '"ND"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFHxA_Result_Num > 0 And PFHxA_Result_Num <= 400000")
arcpy.management.CalculateField(targetTable, "PFHxA_NDE", '"D"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFHxA_Result_Num > 400000")
arcpy.management.CalculateField(targetTable, "PFHxA_NDE", '"E"')


# PFHxS
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFHxS_Result_Num = 0")
arcpy.management.CalculateField(targetTable, "PFHxS_NDE", '"ND"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFHxS_Result_Num > 0 And PFHxS_Result_Num <= 51")
arcpy.management.CalculateField(targetTable, "PFHxS_NDE", '"D"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFHxS_Result_Num > 51")
arcpy.management.CalculateField(targetTable, "PFHxS_NDE", '"E"')


# PFOA
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFOA_Result_Num = 0")
arcpy.management.CalculateField(targetTable, "PFOA_NDE", '"ND"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFOA_Result_Num > 0 And PFOA_Result_Num <= 8")
arcpy.management.CalculateField(targetTable, "PFOA_NDE", '"D"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFOA_Result_Num > 8")
arcpy.management.CalculateField(targetTable, "PFOA_NDE", '"E"')


# PFOS
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFOS_Result_Num = 0")
arcpy.management.CalculateField(targetTable, "PFOS_NDE", '"ND"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFOS_Result_Num > 0 And PFOS_Result_Num <= 8")
arcpy.management.CalculateField(targetTable, "PFOS_NDE", '"D"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFOS_Result_Num > 8")
arcpy.management.CalculateField(targetTable, "PFOS_NDE", '"E"')


# PFNA
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFNA_Result_Num = 0")
arcpy.management.CalculateField(targetTable, "PFNA_NDE", '"ND"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFNA_Result_Num > 0 And PFNA_Result_Num <= 6")
arcpy.management.CalculateField(targetTable, "PFNA_NDE", '"D"')
arcpy.management.SelectLayerByAttribute(targetTable, "NEW_SELECTION", "PFNA_Result_Num > 6")
arcpy.management.CalculateField(targetTable, "PFNA_NDE", '"E"')

arcpy.management.SelectLayerByAttribute(targetTable, "CLEAR_SELECTION", "CLEAR_SELECTION")

# XY Join

arcpy.management.JoinField(targetTable, "AddressID", addresses, "AddressID", "displayx;displayy")

# Make XY events layer

xyEvent = site_ + "_Pivoted" + "_XYEvent"

gcs = '\
GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",SPHEROID["WGS_1984",6378137.0,298.257223563]],PRIMEM["Greenwich",0.0],\
UNIT["Degree",0.0174532925199433]];-400 -400 1000000000;-100000 10000;-100000 10000;8.98315284119521E-09;0.001;0.001;IsHighPrecision'

arcpy.management.MakeXYEventLayer(targetTable, "displayx", "displayy", xyEvent, gcs)

# Seeing an address that was in the all results flat file but didn't make it to this XY event layer?
# Check if the address was only sampled post-filter. In Grayling for example, there is 1 address like this.

# Create a feature class from the XY events layer

inputXY = m.listLayers(xyEvent)[0]
outLoc = location + "/" + gdb
output = site_ + "_Pivoted_FC"

field_mapping = '\
Site "Site" true true false 255 Text 0 0,First,#,inputXY,Site,0,255;\
AddressID "Address ID" true true false 10 Text 0 0,First,#,inputXY,AddressID,0,10;\
Site_Name "Site Name" true true false 255 Text 0 0,First,#,inputXY,Site_Name,0,255;\
Site_Subarea "Site Subarea" true true false 255 Text 0 0,First,#,inputXY,Site_Subarea,0,255;\
Data_File_Name "Data File Name" true true false 255 Text 0 0,First,#,inputXY,Data_File_Name,0,255;\
Report_File_Name "Report File Name" true true false 255 Text 0 0,First,#,inputXY,Report_File_Name,0,255;\
Lab_Name "Lab Name" true true false 255 Text 0 0,First,#,inputXY,Lab_Name,0,255;\
Lab_Work_Order "Lab Work Order" true true false 255 Text 0 0,First,#,inputXY,Lab_Work_Order,0,255;\
Lab_Sample_ID "Lab Sample ID" true true false 255 Text 0 0,First,#,inputXY,Lab_Sample_ID,0,255;\
Field_Sample_ID "Field Sample ID" true true false 255 Text 0 0,First,#,inputXY,Field_Sample_ID,0,255;\
Field_Location_Code "Field Location Code" true true false 255 Text 0 0,First,#,inputXY,Field_Location_Code,0,255;\
Duplicate "Duplicate" true true false 255 Text 0 0,First,#,inputXY,Duplicate,0,255;\
Sampled_Address_Clean "Sampled Address Clean" true true false 255 Text 0 0,First,#,inputXY,Sampled_Address_Clean,0,255;\
Sampling_Round "Sampling Round" true true false 255 Text 0 0,First,#,inputXY,Sampling_Round,0,255;\
Sample_PrePost "Sample PrePost" true true false 255 Text 0 0,First,#,inputXY,Sample_PrePost,0,255;\
Collect_Date "Collect Date" true true false 8 Date 0 0,First,#,inputXY,Collect_Date,-1,-1;\
Collected_By "Collected By" true true false 255 Text 0 0,First,#,inputXY,Collected_By,0,255;\
Matrix_Stdz "Matrix" true true false 255 Text 0 0,First,#,inputXY,Matrix_Stdz,0,255;\
Analysis_Method_Stdz "Analysis Method" true true false 255 Text 0 0,First,#,inputXY,Analysis_Method_Stdz,0,255;\
Sample_NDE "Sample NDE" true true false 2 Text 0 0,First,#,inputXY,Sample_NDE,0,2;\
Sample_TotalPFAS "Sample Total PFAS" true true false 8 Double 0 0,First,#,inputXY,Sample_TotalPFAS,-1,-1;\
HFPO_DA_GenX_NDE "HFPO-DA (GenX) NDE" true true false 2 Text 0 0,First,#,inputXY,HFPO_DA_GenX_NDE,0,2;\
HFPO_DA_GenX_Result_Num "HFPO-DA (GenX) Result Num" true true false 8 Double 0 0,First,#,inputXY,HFPO_DA_GenX_Result_Num,-1,-1;\
HFPO_DA_GenX_Result_Qualifier "HFPO-DA (GenX) Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,HFPO_DA_GenX_Result_Qualifier,0,20;\
PFBS_NDE "PFBS NDE" true true false 2 Text 0 0,First,#,inputXY,PFBS_NDE,0,2;\
PFBS_Result_Num "PFBS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFBS_Result_Num,-1,-1;\
PFBS_Result_Qualifier "PFBS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFBS_Result_Qualifier,0,20;\
PFHxA_NDE "PFHxA NDE" true true false 2 Text 0 0,First,#,inputXY,PFHxA_NDE,0,2;\
PFHxA_Result_Num "PFHxA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFHxA_Result_Num,-1,-1;\
PFHxA_Result_Qualifier "PFHxA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFHxA_Result_Qualifier,0,20;\
PFHxS_NDE "PFHxS NDE" true true false 2 Text 0 0,First,#,inputXY,PFHxS_NDE,0,2;\
PFHxS_Result_Num "PFHxS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFHxS_Result_Num,-1,-1;\
PFHxS_Result_Qualifier "PFHxS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFHxS_Result_Qualifier,0,20;\
PFOA_NDE "PFOA NDE" true true false 2 Text 0 0,First,#,inputXY,PFOA_NDE,0,2;\
PFOA_Result_Num "PFOA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFOA_Result_Num,-1,-1;\
PFOA_Result_Qualifier "PFOA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFOA_Result_Qualifier,0,20;\
PFOS_NDE "PFOS NDE" true true false 2 Text 0 0,First,#,inputXY,PFOS_NDE,0,2;\
PFOS_Result_Num "PFOS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFOS_Result_Num,-1,-1;\
PFOS_Result_Qualifier "PFOS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFOS_Result_Qualifier,0,20;\
PFNA_NDE "PFNA NDE" true true false 2 Text 0 0,First,#,inputXY,PFNA_NDE,0,2;\
PFNA_Result_Num "PFNA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFNA_Result_Num,-1,-1;\
PFNA_Result_Qualifier "PFNA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFNA_Result_Qualifier,0,20;\
ADONA_Result_Num "ADONA Result Num" true true false 8 Double 0 0,First,#,inputXY,ADONA_Result_Num,-1,-1;\
ADONA_Result_Qualifier "ADONA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,ADONA_Result_Qualifier,0,20;\
Br_PFHxS_Result_Num "Br-PFHxS Result Num" true true false 8 Double 0 0,First,#,inputXY,Br_PFHxS_Result_Num,-1,-1;\
Br_PFHxS_Result_Qualifier "Br-PFHxS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,Br_PFHxS_Result_Qualifier,0,20;\
Br_PFOA_Result_Num "Br-PFOA Result Num" true true false 8 Double 0 0,First,#,inputXY,Br_PFOA_Result_Num,-1,-1;\
Br_PFOA_Result_Qualifier "Br-PFOA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,Br_PFOA_Result_Qualifier,0,20;\
Br_PFOS_Result_Num "Br-PFOS Result Num" true true false 8 Double 0 0,First,#,inputXY,Br_PFOS_Result_Num,-1,-1;\
Br_PFOS_Result_Qualifier "Br-PFOS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,Br_PFOS_Result_Qualifier,0,20;\
EtFOSA_Result_Num "EtFOSA Result Num" true true false 8 Double 0 0,First,#,inputXY,EtFOSA_Result_Num,-1,-1;\
EtFOSA_Result_Qualifier "EtFOSA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,EtFOSA_Result_Qualifier,0,20;\
EtFOSAA_Result_Num "EtFOSAA Result Num" true true false 8 Double 0 0,First,#,inputXY,EtFOSAA_Result_Num,-1,-1;\
EtFOSAA_Result_Qualifier "EtFOSAA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,EtFOSAA_Result_Qualifier,0,20;\
EtFOSE_Result_Num "EtFOSE Result Num" true true false 8 Double 0 0,First,#,inputXY,EtFOSE_Result_Num,-1,-1;\
EtFOSE_Result_Qualifier "EtFOSE Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,EtFOSE_Result_Qualifier,0,20;\
F_11Cl_PF3OUdS_Result_Num "11Cl-PF3OUdS Result Num" true true false 8 Double 0 0,First,#,inputXY,F_11Cl_PF3OUdS_Result_Num,-1,-1;\
F_11Cl_PF3OUdS_Result_Qualifier "11Cl-PF3OUdS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,F_11Cl_PF3OUdS_Result_Qualifier,0,20;\
F_3_3FTCA_Result_Num "3:3 FTCA Result Num" true true false 8 Double 0 0,First,#,inputXY,F_3_3FTCA_Result_Num,-1,-1;\
F_3_3FTCA_Result_Qualifier "3:3 FTCA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,F_3_3FTCA_Result_Qualifier,0,20;\
F_4_2FTS_Result_Num "4:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputXY,F_4_2FTS_Result_Num,-1,-1;\
F_4_2FTS_Result_Qualifier "4:2 FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,F_4_2FTS_Result_Qualifier,0,20;\
F_5_3FTCA_Result_Num "5:3 FTCA Result Num" true true false 8 Double 0 0,First,#,inputXY,F_5_3FTCA_Result_Num,-1,-1;\
F_5_3FTCA_Result_Qualifier "5:3 FTCA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,F_5_3FTCA_Result_Qualifier,0,20;\
F_6_2FTS_Result_Num "6:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputXY,F_6_2FTS_Result_Num,-1,-1;\
F_6_2FTS_Result_Qualifier "6:2 FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,F_6_2FTS_Result_Qualifier,0,20;\
F_7_3FTCA_Result_Num "7:3 FTCA Result Num" true true false 8 Double 0 0,First,#,inputXY,F_7_3FTCA_Result_Num,-1,-1;\
F_7_3FTCA_Result_Qualifier "7:3 FTCA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,F_7_3FTCA_Result_Qualifier,0,20;\
F_8_2FTS_Result_Num "8:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputXY,F_8_2FTS_Result_Num,-1,-1;\
F_8_2FTS_Result_Qualifier "8:2 FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,F_8_2FTS_Result_Qualifier,0,20;\
F_9Cl_PF3ONS_Result_Num "9Cl-PF3ONS Result Num" true true false 8 Double 0 0,First,#,inputXY,F_9Cl_PF3ONS_Result_Num,-1,-1;\
F_9Cl_PF3ONS_Result_Qualifier "9Cl-PF3ONS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,F_9Cl_PF3ONS_Result_Qualifier,0,20;\
L_6_2FTS_Result_Num "L-6:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputXY,L_6_2FTS_Result_Num,-1,-1;\
L_6_2FTS_Result_Qualifier "L-6:2 FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_6_2FTS_Result_Qualifier,0,20;\
L_8_2FTS_Result_Num "L-8:2 FTS Result Num" true true false 8 Double 0 0,First,#,inputXY,L_8_2FTS_Result_Num,-1,-1;\
L_8_2FTS_Result_Qualifier "L-8:2FTS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_8_2FTS_Result_Qualifier,0,20;\
L_PFBA_Result_Num "L-PFBA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFBA_Result_Num,-1,-1;\
L_PFBA_Result_Qualifier "L-PFBA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFBA_Result_Qualifier,0,20;\
L_PFBS_Result_Num "L-PFBS Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFBS_Result_Num,-1,-1;\
L_PFBS_Result_Qualifier "L-PFBS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFBS_Result_Qualifier,0,20;\
L_PFDA_Result_Num "L-PFDA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFDA_Result_Num,-1,-1;\
L_PFDA_Result_Qualifier "L-PFDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFDA_Result_Qualifier,0,20;\
L_PFDoDA_Result_Num "L-PFDoDA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFDoDA_Result_Num,-1,-1;\
L_PFDoDA_Result_Qualifier "L-PFDoDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFDoDA_Result_Qualifier,0,20;\
L_PFDS_Result_Num "L-PFDS Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFDS_Result_Num,-1,-1;\
L_PFDS_Result_Qualifier "L-PFDS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFDS_Result_Qualifier,0,20;\
L_PFHpA_Result_Num "L-PFHpA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFHpA_Result_Num,-1,-1;\
L_PFHpA_Result_Qualifier "L-PFHpA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFHpA_Result_Qualifier,0,20;\
L_PFHpS_Result_Num "L-PFHpS Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFHpS_Result_Num,-1,-1;\
L_PFHpS_Result_Qualifier "L-PFHpS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFHpS_Result_Qualifier,0,20;\
L_PFHxA_Result_Num "L-PFHxA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFHxA_Result_Num,-1,-1;\
L_PFHxA_Result_Qualifier "L-PFHxA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFHxA_Result_Qualifier,0,20;\
L_PFHxDA_Result_Num "L-PFHxDA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFHxDA_Result_Num,-1,-1;\
L_PFHxDA_Result_Qualifier "L-PFHxDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFHxDA_Result_Qualifier,0,20;\
L_PFHxS_Result_Num "L-PFHxS Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFHxS_Result_Num,-1,-1;\
L_PFHxS_Result_Qualifier "L-PFHxS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFHxS_Result_Qualifier,0,20;\
L_PFNA_Result_Num "L-PFNA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFNA_Result_Num,-1,-1;\
L_PFNA_Result_Qualifier "L-PFNA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFNA_Result_Qualifier,0,20;\
L_PFOA_Result_Num "L-PFOA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFOA_Result_Num,-1,-1;\
L_PFOA_Result_Qualifier "L-PFOA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFOA_Result_Qualifier,0,20;\
L_PFODA_Result_Num "L-PFODA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFODA_Result_Num,-1,-1;\
L_PFODA_Result_Qualifier "L-PFODA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFODA_Result_Qualifier,0,20;\
L_PFOS_Result_Num "L-PFOS Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFOS_Result_Num,-1,-1;\
L_PFOS_Result_Qualifier "L-PFOS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFOS_Result_Qualifier,0,20;\
L_PFOSA_Result_Num "L-PFOSA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFOSA_Result_Num,-1,-1;\
L_PFOSA_Result_Qualifier "L-PFOSA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFOSA_Result_Qualifier,0,20;\
L_PFPeA_Result_Num "L-PFPeA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFPeA_Result_Num,-1,-1;\
L_PFPeA_Result_Qualifier "L-PFPeA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFPeA_Result_Qualifier,0,20;\
L_PFTeDA_Result_Num "L-PFTeDA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFTeDA_Result_Num,-1,-1;\
L_PFTeDA_Result_Qualifier "L-PFTeDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFTeDA_Result_Qualifier,0,20;\
L_PFTrDA_Result_Num "L-PFTrDA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFTrDA_Result_Num,-1,-1;\
L_PFTrDA_Result_Qualifier "L-PFTrDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFTrDA_Result_Qualifier,0,20;\
L_PFUnDA_Result_Num "L-PFUnDA Result Num" true true false 8 Double 0 0,First,#,inputXY,L_PFUnDA_Result_Num,-1,-1;\
L_PFUnDA_Result_Qualifier "L-PFUnDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,L_PFUnDA_Result_Qualifier,0,20;\
MeFOSA_Result_Num "MeFOSA Result Num" true true false 8 Double 0 0,First,#,inputXY,MeFOSA_Result_Num,-1,-1;\
MeFOSA_Result_Qualifier "MeFOSA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,MeFOSA_Result_Qualifier,0,20;\
MeFOSAA_Result_Num "MeFOSAA Result Num" true true false 8 Double 0 0,First,#,inputXY,MeFOSAA_Result_Num,-1,-1;\
MeFOSAA_Result_Qualifier "MeFOSAA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,MeFOSAA_Result_Qualifier,0,20;\
MeFOSE_Result_Num "MeFOSE Result Num" true true false 8 Double 0 0,First,#,inputXY,MeFOSE_Result_Num,-1,-1;\
MeFOSE_Result_Qualifier "MeFOSE Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,MeFOSE_Result_Qualifier,0,20;\
NFDHA_Result_Num "NFDHA Result Num" true true false 8 Double 0 0,First,#,inputXY,NFDHA_Result_Num,-1,-1;\
NFDHA_Result_Qualifier "NFDHA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,NFDHA_Result_Qualifier,0,20;\
PFBA_Result_Num "PFBA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFBA_Result_Num,-1,-1;\
PFBA_Result_Qualifier "PFBA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFBA_Result_Qualifier,0,20;\
PFBSA_Result_Num "PFBSA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFBSA_Result_Num,-1,-1;\
PFBSA_Result_Qualifier "PFBSA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFBSA_Result_Qualifier,0,20;\
PFDA_Result_Num "PFDA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFDA_Result_Num,-1,-1;\
PFDA_Result_Qualifier "PFDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFDA_Result_Qualifier,0,20;\
PFDoDA_Result_Num "PFDoDA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFDoDA_Result_Num,-1,-1;\
PFDoDA_Result_Qualifier "PFDoDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFDoDA_Result_Qualifier,0,20;\
PFDS_Result_Num "PFDS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFDS_Result_Num,-1,-1;\
PFDS_Result_Qualifier "PFDS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFDS_Result_Qualifier,0,20;\
PFECHS_Result_Num "PFECHS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFECHS_Result_Num,-1,-1;\
PFECHS_Result_Qualifier "PFECHS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFECHS_Result_Qualifier,0,20;\
PFEESA_Result_Num "PFEESA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFEESA_Result_Num,-1,-1;\
PFEESA_Result_Qualifier "PFEESA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFEESA_Result_Qualifier,0,20;\
PFHpA_Result_Num "PFHpA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFHpA_Result_Num,-1,-1;\
PFHpA_Result_Qualifier "PFHpA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFHpA_Result_Qualifier,0,20;\
PFHpS_Result_Num "PFHpS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFHpS_Result_Num,-1,-1;\
PFHpS_Result_Qualifier "PFHpS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFHpS_Result_Qualifier,0,20;\
PFHxDA_Result_Num "PFHxDA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFHxDA_Result_Num,-1,-1;\
PFHxDA_Result_Qualifier "PFHxDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFHxDA_Result_Qualifier,0,20;\
PFHxSA_Result_Num "PFHxSA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFHxSA_Result_Num,-1,-1;\
PFHxSA_Result_Qualifier "PFHxSA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFHxSA_Result_Qualifier,0,20;\
PFMBA_Result_Num "PFMBA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFMBA_Result_Num,-1,-1;\
PFMBA_Result_Qualifier "PFMBA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFMBA_Result_Qualifier,0,20;\
PFMPA_Result_Num "PFMPA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFMPA_Result_Num,-1,-1;\
PFMPA_Result_Qualifier "PFMPA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFMPA_Result_Qualifier,0,20;\
PFNS_Result_Num "PFNS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFNS_Result_Num,-1,-1;\
PFNS_Result_Qualifier "PFNS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFNS_Result_Qualifier,0,20;\
PFOA_PFOS_Result_Num "PFOA+PFOS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFOA_PFOS_Result_Num,-1,-1;\
PFOA_PFOS_Result_Qualifier "PFOA+PFOS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFOA_PFOS_Result_Qualifier,0,20;\
PFODA_Result_Num "PFODA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFODA_Result_Num,-1,-1;\
PFODA_Result_Qualifier "PFODA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFODA_Result_Qualifier,0,20;\
PFOSA_Result_Num "PFOSA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFOSA_Result_Num,-1,-1;\
PFOSA_Result_Qualifier "PFOSA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFOSA_Result_Qualifier,0,20;\
PFPeA_Result_Num "PFPeA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFPeA_Result_Num,-1,-1;\
PFPeA_Result_Qualifier "PFPeA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFPeA_Result_Qualifier,0,20;\
PFPeS_Result_Num "PFPeS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFPeS_Result_Num,-1,-1;\
PFPeS_Result_Qualifier "PFPeS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFPeS_Result_Qualifier,0,20;\
PFPrS_Result_Num "PFPrS Result Num" true true false 8 Double 0 0,First,#,inputXY,PFPrS_Result_Num,-1,-1;\
PFPrS_Result_Qualifier "PFPrS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFPrS_Result_Qualifier,0,20;\
PFTeDA_Result_Num "PFTeDA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFTeDA_Result_Num,-1,-1;\
PFTeDA_Result_Qualifier "PFTeDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFTeDA_Result_Qualifier,0,20;\
PFTrDA_Result_Num "PFTrDA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFTrDA_Result_Num,-1,-1;\
PFTrDA_Result_Qualifier "PFTrDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFTrDA_Result_Qualifier,0,20;\
PFUnDA_Result_Num "PFUnDA Result Num" true true false 8 Double 0 0,First,#,inputXY,PFUnDA_Result_Num,-1,-1;\
PFUnDA_Result_Qualifier "PFUnDA Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,PFUnDA_Result_Qualifier,0,20;\
TotalPFAS_Result_Num "Total PFAS Result Num" true true false 8 Double 0 0,First,#,inputXY,TotalPFAS_Result_Num,-1,-1;\
TotalPFAS_Result_Qualifier "Total PFAS Result Qualifier" true true false 20 Text 0 0,First,#,inputXY,TotalPFAS_Result_Qualifier,0,20'

arcpy.conversion.FeatureClassToFeatureClass(inputXY, outLoc, output, '', field_mapping)

# Append FC to PFAS sampling results layer on Portal
# 
